import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from config import DATA_DIR, EXCEL_PATH

HEADERS = [
    "Date & Time",
    "Service Name",
    "Amount",
    "Customer Name",
    "Mobile",
    "Aadhaar",
]


def validate_submission(name, mobile, aadhaar):
    errors = []
    name = (name or "").strip()
    mobile = (mobile or "").strip()
    aadhaar = (aadhaar or "").strip()

    if len(name) < 2:
        errors.append("Customer name must be at least 2 characters.")
    if not re.fullmatch(r"\d{10}", mobile):
        errors.append("Mobile number must be exactly 10 digits.")
    if not re.fullmatch(r"\d{12}", aadhaar):
        errors.append("Aadhaar number must be exactly 12 digits.")

    return errors, name, mobile, aadhaar


def _ensure_workbook():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    if not EXCEL_PATH.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Records"
        ws.append(HEADERS)
        wb.save(EXCEL_PATH)


def append_record(service_name, amount, customer_name, mobile, aadhaar):
    _ensure_workbook()
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        service_name,
        amount,
        customer_name,
        mobile,
        aadhaar,
    ])
    wb.save(EXCEL_PATH)
